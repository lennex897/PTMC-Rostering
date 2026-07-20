from collections import Counter
from pathlib import Path

import openpyxl

from roster_engine.roster_grid import parse_roster_grid


WORKBOOK_PATH = Path(
    "reference/Scheduling Roster 2026.xlsx"
)
WORKSHEET_NAME = "May-2026-Roster"


grid = parse_roster_grid(
    workbook_path=WORKBOOK_PATH,
    worksheet_name=WORKSHEET_NAME,
)

workbook = openpyxl.load_workbook(
    WORKBOOK_PATH,
    read_only=False,
    data_only=True,
)

try:
    worksheet = workbook[WORKSHEET_NAME]

    values = Counter()
    shown = 0

    day_by_column = {
        day.column_number: day
        for day in grid.days
    }

    personnel_by_row = {
        row.row_number: row.label
        for row in grid.personnel_rows
    }

    first_date_column = min(day_by_column)
    last_date_column = max(day_by_column)

    print(f"Worksheet: {WORKSHEET_NAME}")
    print(f"Calendar days: {len(grid.days)}")
    print(f"Personnel rows: {len(grid.personnel_rows)}")
    print()
    print("First 50 non-empty cells:")

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=5,
            max_row=worksheet.max_row,
            min_col=first_date_column,
            max_col=last_date_column,
            values_only=True,
        ),
        start=5,
    ):
        person_label = personnel_by_row.get(row_number)

        if not person_label:
            continue

        for offset, raw_value in enumerate(row):
            if raw_value is None:
                continue

            value = str(raw_value).strip()

            if not value:
                continue

            column_number = first_date_column + offset
            roster_day = day_by_column.get(column_number)

            if roster_day is None:
                continue

            values[value] += 1

            if shown < 50:
                print(
                    f"{roster_day.roster_date} | "
                    f"{person_label} | "
                    f"{value!r}"
                )
                shown += 1

    print()
    print("Unique assignment values:")

    for value, count in values.most_common():
        print(f"{value!r}: {count}")

finally:
    workbook.close()
