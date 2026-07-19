from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


# Change this to the local test copy of your leave workbook.
LEAVE_WORKBOOK_PATH = Path("reference/Leave_Workbook.xlsx")

HEADER_KEYWORDS = {
    "name",
    "rank",
    "personnel",
    "date",
    "day",
    "leave",
    "start",
    "end",
    "from",
    "to",
    "type",
    "status",
    "remarks",
    "reason",
    "mc",
    "off",
    "course",
    "ord",
    "unit",
    "section",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def is_possible_header(value: str) -> bool:
    normalised = value.lower().replace("_", " ").strip()

    return any(
        keyword == normalised
        or keyword in normalised.split()
        for keyword in HEADER_KEYWORDS
    )


if not LEAVE_WORKBOOK_PATH.exists():
    raise FileNotFoundError(
        f"Leave workbook not found: {LEAVE_WORKBOOK_PATH.resolve()}\n"
        "Update LEAVE_WORKBOOK_PATH to point to your local test workbook."
    )

workbook = openpyxl.load_workbook(
    LEAVE_WORKBOOK_PATH,
    read_only=True,
    data_only=True,
)

try:
    print("\nAVAILABLE WORKSHEETS")
    print("--------------------")

    for sheet_name in workbook.sheetnames:
        print(sheet_name)

    print("\nWORKSHEET STRUCTURE")
    print("-------------------")

    for worksheet in workbook.worksheets:
        print(f"\nWorksheet: {worksheet.title}")
        print(
            f"Rows: {worksheet.max_row}, "
            f"Columns: {worksheet.max_column}"
        )

        candidates_found = 0

        # Inspect only the first 30 rows for likely headings.
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(30, worksheet.max_row),
                values_only=True,
            ),
            start=1,
        ):
            matched_cells: list[str] = []

            for column_number, raw_value in enumerate(row, start=1):
                value = clean_text(raw_value)

                if not value or not is_possible_header(value):
                    continue

                column_letter = get_column_letter(column_number)
                matched_cells.append(
                    f"{column_letter}={value!r}"
                )

            if matched_cells:
                candidates_found += 1
                print(
                    f"Possible header row {row_number}: "
                    + " | ".join(matched_cells)
                )

        if candidates_found == 0:
            print(
                "No obvious header keywords were found "
                "in the first 30 rows."
            )

finally:
    workbook.close()