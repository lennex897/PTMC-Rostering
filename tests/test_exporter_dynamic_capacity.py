from openpyxl import Workbook

from roster_engine.exporter import (
    BASE_PERSONNEL_CAPACITY,
    _expand_personnel_blocks,
)


def test_original_capacities() -> None:
    assert BASE_PERSONNEL_CAPACITY["PT"] == 37
    assert BASE_PERSONNEL_CAPACITY["RH"] == 20


def test_pt_expands_to_38_people() -> None:
    workbook = Workbook()
    worksheet = workbook.active

    for row in range(5, 72):
        worksheet.cell(row=row, column=2).value = f"ROW {row}"

    blocks = _expand_personnel_blocks(
        worksheet,
        pt_count=38,
        rh_count=20,
    )

    assert len(list(blocks["PT"])) == 38
    assert len(list(blocks["RH"])) == 20
    assert blocks["RH"].start == 53


def test_rh_can_expand_independently() -> None:
    workbook = Workbook()
    worksheet = workbook.active

    blocks = _expand_personnel_blocks(
        worksheet,
        pt_count=37,
        rh_count=22,
    )

    assert len(list(blocks["PT"])) == 37
    assert len(list(blocks["RH"])) == 22
