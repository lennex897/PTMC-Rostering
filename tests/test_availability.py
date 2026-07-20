from datetime import date
from pathlib import Path

import openpyxl

from roster_engine.availability import load_availability


def test_load_availability(tmp_path: Path) -> None:
    workbook_path = tmp_path / "Leave_Off_Impt Dates Forecast.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Aug 26"

    worksheet["A1"] = "Name"
    worksheet["B1"] = date(2026, 8, 1)
    worksheet["C1"] = date(2026, 8, 2)

    worksheet["A2"] = "CPL TEST PERSON"
    worksheet["B2"] = "AL"
    worksheet["C2"] = ""

    worksheet["A3"] = "LCP SECOND PERSON"
    worksheet["B3"] = ""
    worksheet["C3"] = "MC"

    workbook.save(workbook_path)

    entries = load_availability(
        workbook_path,
        "Aug 26",
    )

    assert len(entries) == 2

    assert entries[0].person_name == "CPL TEST PERSON"
    assert entries[0].unavailable_date == date(2026, 8, 1)
    assert entries[0].reason == "AL"

    assert entries[1].person_name == "LCP SECOND PERSON"
    assert entries[1].unavailable_date == date(2026, 8, 2)
    assert entries[1].reason == "MC"


def test_missing_worksheet_raises_error(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "leave.xlsx"

    workbook = openpyxl.Workbook()
    workbook.save(workbook_path)

    try:
        load_availability(
            workbook_path,
            "Aug 26",
        )
    except ValueError as error:
        assert "Aug 26" in str(error)
    else:
        raise AssertionError(
            "Expected ValueError for missing worksheet."
        )

def test_parse_month_year_from_sheet_name() -> None:
    from roster_engine.availability import (
        parse_month_year_from_sheet_name,
    )

    assert parse_month_year_from_sheet_name(
        "Aug 26"
    ) == (2026, 8)

    assert parse_month_year_from_sheet_name(
        "Jul-2026"
    ) == (2026, 7)

    assert parse_month_year_from_sheet_name(
        "August 2026"
    ) == (2026, 8)

    assert parse_month_year_from_sheet_name(
        "Base"
    ) is None


def test_load_availability_with_numbered_day_headers(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "leave.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Aug 26"

    worksheet["B2"] = "Name/Date"
    worksheet["C2"] = 1
    worksheet["D2"] = 2
    worksheet["E2"] = 3

    worksheet["B3"] = "PERSON A"
    worksheet["C3"] = "AL"
    worksheet["D3"] = ""
    worksheet["E3"] = "AM MA PM AL"

    worksheet["B4"] = "PERSON B"
    worksheet["C4"] = "HL"
    worksheet["D4"] = "ORD"

    workbook.save(workbook_path)

    entries = load_availability(
        workbook_path=workbook_path,
        worksheet_name="Aug 26",
    )

    assert {
        (
            entry.person_name,
            entry.unavailable_date,
            entry.reason,
        )
        for entry in entries
    } == {
        (
            "PERSON A",
            date(2026, 8, 1),
            "AL",
        ),
        (
            "PERSON A",
            date(2026, 8, 3),
            "AL",
        ),
        (
            "PERSON B",
            date(2026, 8, 1),
            "HL",
        ),
        (
            "PERSON B",
            date(2026, 8, 2),
            "ORD",
        ),
    }
