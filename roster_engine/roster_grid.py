from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


EVENT_ROW = 2
DATE_ROW = 3
DAY_ROW = 4

PERSONNEL_START_ROW = 5
PERSONNEL_COLUMN = 2

SCHEDULE_START_COLUMN = 3


@dataclass(frozen=True)
class RosterDay:
    column_number: int
    roster_date: date
    day_name: str
    event: str


@dataclass(frozen=True)
class PersonnelRow:
    row_number: int
    label: str


@dataclass(frozen=True)
class RosterGrid:
    worksheet_name: str
    days: list[RosterDay]
    personnel_rows: list[PersonnelRow]


def clean_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def normalise_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    return None


def find_roster_days(worksheet: Worksheet) -> list[RosterDay]:
    days: list[RosterDay] = []
    date_sequence_started = False

    for column_number in range(
        SCHEDULE_START_COLUMN,
        worksheet.max_column + 1,
    ):
        raw_date = worksheet.cell(
            row=DATE_ROW,
            column=column_number,
        ).value

        roster_date = normalise_date(raw_date)

        if roster_date is None:
            if date_sequence_started:
                # The calendar sequence has ended. Remaining columns
                # contain summary fields such as DM, CS1 and totals.
                break

            # Allow blank or label columns before the first calendar day.
            continue

        date_sequence_started = True

        event = clean_text(
            worksheet.cell(
                row=EVENT_ROW,
                column=column_number,
            ).value
        )

        day_name = clean_text(
            worksheet.cell(
                row=DAY_ROW,
                column=column_number,
            ).value
        )

        # Fall back to the actual date if cached formula results
        # are unavailable.
        if not day_name:
            day_name = roster_date.strftime("%a")

        days.append(
            RosterDay(
                column_number=column_number,
                roster_date=roster_date,
                day_name=day_name,
                event=event,
            )
        )

    return days


def find_personnel_rows(
    worksheet: Worksheet,
) -> list[PersonnelRow]:
    personnel_rows: list[PersonnelRow] = []

    for row_number in range(
        PERSONNEL_START_ROW,
        worksheet.max_row + 1,
    ):
        label = clean_text(
            worksheet.cell(
                row=row_number,
                column=PERSONNEL_COLUMN,
            ).value
        )

        if not label:
            continue

        personnel_rows.append(
            PersonnelRow(
                row_number=row_number,
                label=label,
            )
        )

    return personnel_rows


def parse_roster_grid(
    workbook_path: Path,
    worksheet_name: str,
) -> RosterGrid:
    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Roster workbook not found: {workbook_path}"
        )

    workbook = openpyxl.load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    try:
        if worksheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet '{worksheet_name}' was not found. "
                f"Available worksheets: {workbook.sheetnames}"
            )

        worksheet = workbook[worksheet_name]

        days = find_roster_days(worksheet)
        personnel_rows = find_personnel_rows(worksheet)

        if not days:
            raise ValueError(
                f"No calendar dates were found in row {DATE_ROW} "
                f"of worksheet '{worksheet_name}'."
            )

        if not personnel_rows:
            raise ValueError(
                f"No personnel rows were found in column B "
                f"from row {PERSONNEL_START_ROW} onward."
            )

        return RosterGrid(
            worksheet_name=worksheet_name,
            days=days,
            personnel_rows=personnel_rows,
        )

    finally:
        workbook.close()