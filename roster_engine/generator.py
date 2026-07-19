from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import openpyxl


@dataclass
class GenerationRequest:
    roster_month: date
    scheduling_roster_path: Path
    leave_workbook_path: Path
    leave_sheet: str
    rulebook_text: str
    assumptions_text: str
    output_path: Path


@dataclass
class GenerationResult:
    success: bool
    output_path: Path | None = None
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    statistics: dict[str, object] = field(default_factory=dict)


def generate_roster(
    request: GenerationRequest,
) -> GenerationResult:
    errors: list[str] = []
    warnings: list[str] = []
    statistics: dict[str, object] = {}

    if not request.scheduling_roster_path.exists():
        errors.append(
            "The scheduling roster reference workbook is missing."
        )

    if not request.leave_workbook_path.exists():
        errors.append(
            "The uploaded leave workbook could not be found."
        )

    if not request.rulebook_text.strip():
        errors.append("The Rulebook is empty.")

    if not request.assumptions_text.strip():
        errors.append("The assumptions document is empty.")

    if errors:
        return GenerationResult(
            success=False,
            errors=errors,
        )

    try:
        scheduling_workbook = openpyxl.load_workbook(
            request.scheduling_roster_path,
            read_only=True,
            data_only=True,
        )

        statistics["scheduling_sheets"] = (
            scheduling_workbook.sheetnames
        )

    except Exception as exc:
        errors.append(
            f"Unable to open scheduling roster: {exc}"
        )

    try:
        leave_workbook = openpyxl.load_workbook(
            request.leave_workbook_path,
            read_only=True,
            data_only=True,
        )

        if request.leave_sheet not in leave_workbook.sheetnames:
            errors.append(
                f"Leave worksheet '{request.leave_sheet}' "
                "could not be found."
            )
        else:
            statistics["leave_sheet"] = request.leave_sheet
            statistics["leave_sheets"] = leave_workbook.sheetnames

    except Exception as exc:
        errors.append(
            f"Unable to open leave workbook: {exc}"
        )

    if errors:
        return GenerationResult(
            success=False,
            errors=errors,
            warnings=warnings,
            statistics=statistics,
        )

    warnings.append(
        "Inputs were validated, but the scheduling algorithm "
        "has not yet been connected."
    )

    return GenerationResult(
        success=False,
        errors=[],
        warnings=warnings,
        statistics=statistics,
    )