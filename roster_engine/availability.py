from __future__ import annotations

import re
from calendar import month_abbr, month_name
from datetime import date, datetime
from pathlib import Path

import openpyxl

from roster_engine.models import AvailabilityEntry
from roster_engine.personnel import normalise_text




BLOCKING_PRIORITY = (
    "AL",
    "ORD",
    "OWADIO",
    "HL",
    "MC",
    "OIL",
    "OFF",
    "MA",
    "AM",
    "PM",
)

BLOCKING_CODES = set(BLOCKING_PRIORITY)


def parse_date_header(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    return None


def parse_month_year_from_sheet_name(
    worksheet_name: str,
) -> tuple[int, int] | None:
    """
    Parse worksheet names such as:

    Aug 26
    Aug 2026
    August 26
    Jul-2026

    Two-digit years are interpreted as 2000–2099.
    """
    normalised_name = normalise_text(
        worksheet_name
    ).replace("-", " ")

    match = re.search(
        r"\b([A-Z]+)\s+(\d{2}|\d{4})\b",
        normalised_name,
    )

    if match is None:
        return None

    month_text = match.group(1)
    year_text = match.group(2)

    month_lookup: dict[str, int] = {}

    for month_number in range(1, 13):
        month_lookup[
            month_abbr[month_number].upper()
        ] = month_number

        month_lookup[
            month_name[month_number].upper()
        ] = month_number

    month_number = month_lookup.get(month_text)

    if month_number is None:
        return None

    year_number = int(year_text)

    if year_number < 100:
        year_number += 2000

    return year_number, month_number


def find_name_header(
    worksheet,
    maximum_header_rows: int = 10,
) -> tuple[int, int] | None:
    """
    Return the row and column containing a recognised personnel
    name header.
    """
    recognised_headers = {
        "NAME/DATE",
        "NAME",
        "PERSONNEL",
    }

    for row_number in range(
        1,
        min(maximum_header_rows, worksheet.max_row) + 1,
    ):
        for column_number in range(
            1,
            worksheet.max_column + 1,
        ):
            value = normalise_text(
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value
            )

            if value in recognised_headers:
                return row_number, column_number

    return None


def build_date_columns(
    worksheet,
    worksheet_name: str,
    header_row: int | None,
) -> dict[int, date]:
    """
    Detect date columns using either:

    1. Full Excel date values.
    2. Day numbers combined with a month and year inferred from
       the worksheet name.
    """
    date_columns: dict[int, date] = {}

    maximum_header_row = min(
        10,
        worksheet.max_row,
    )

    # First preserve support for worksheets containing actual
    # Excel dates.
    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        for row_number in range(
            1,
            maximum_header_row + 1,
        ):
            parsed_date = parse_date_header(
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value
            )

            if parsed_date is not None:
                date_columns[column_number] = parsed_date
                break

    if date_columns:
        return date_columns

    inferred_month = parse_month_year_from_sheet_name(
        worksheet_name
    )

    if inferred_month is None:
        return {}

    year_number, month_number = inferred_month

    candidate_rows: list[int]

    if header_row is not None:
        candidate_rows = [header_row]
    else:
        candidate_rows = list(
            range(1, maximum_header_row + 1)
        )

    for row_number in candidate_rows:
        candidate_columns: dict[int, date] = {}

        for column_number in range(
            1,
            worksheet.max_column + 1,
        ):
            value = worksheet.cell(
                row=row_number,
                column=column_number,
            ).value

            if isinstance(value, bool):
                continue

            if not isinstance(value, int):
                continue

            try:
                parsed_date = date(
                    year_number,
                    month_number,
                    value,
                )
            except ValueError:
                continue

            candidate_columns[column_number] = parsed_date

        # Accept rows that contain a consecutive run of day
        # numbers (e.g. 1,2,3 or 1..31).
        day_numbers = sorted(
            parsed_date.day
            for parsed_date in candidate_columns.values()
        )

        if (
            len(day_numbers) >= 2
            and day_numbers
            == list(
                range(
                    day_numbers[0],
                    day_numbers[-1] + 1,
                )
            )
        ):
            return candidate_columns

    return {}


def extract_blocking_code(
    raw_reason: object,
) -> str | None:
    """
    Return the highest-priority blocking code found in a cell.

    For example:
        "AM MA PM AL" -> "AL"
    """
    reason = normalise_text(raw_reason)

    if not reason:
        return None

    tokens = set(
        re.findall(
            r"[A-Z0-9]+",
            reason,
        )
    )

    present_codes = tokens & BLOCKING_CODES

    for code in BLOCKING_PRIORITY:
        if code in present_codes:
            return code

    return None


def load_availability(
    workbook_path: Path,
    worksheet_name: str,
) -> list[AvailabilityEntry]:
    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Leave workbook not found: {workbook_path}"
        )

    workbook = openpyxl.load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    try:
        if worksheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet '{worksheet_name}' was not found."
            )

        worksheet = workbook[worksheet_name]

        name_header = find_name_header(worksheet)

        if name_header is None:
            header_row = None
            name_column = 1
            personnel_start_row = 1
        else:
            header_row, name_column = name_header
            personnel_start_row = header_row + 1

        date_columns = build_date_columns(
            worksheet=worksheet,
            worksheet_name=worksheet_name,
            header_row=header_row,
        )

        if not date_columns:
            raise ValueError(
                "No date columns were found in the selected "
                "leave worksheet."
            )

        entries: list[AvailabilityEntry] = []

        for row_number in range(
            personnel_start_row,
            worksheet.max_row + 1,
        ):
            person_name = normalise_text(
                worksheet.cell(
                    row=row_number,
                    column=name_column,
                ).value
            )

            if not person_name:
                continue

            for column_number, unavailable_date in (
                date_columns.items()
            ):
                matched_code = extract_blocking_code(
                    worksheet.cell(
                        row=row_number,
                        column=column_number,
                    ).value
                )

                if matched_code is None:
                    continue

                entries.append(
                    AvailabilityEntry(
                        person_name=person_name,
                        unavailable_date=unavailable_date,
                        reason=matched_code,
                    )
                )

        return entries

    finally:
        workbook.close()
