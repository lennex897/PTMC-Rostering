from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from roster_engine.models import Assignment, Person, Schedule
from roster_engine.personnel import normalise_text
from roster_engine.roster_grid import (
    DATE_ROW,
    PERSONNEL_COLUMN,
    PERSONNEL_START_ROW,
    SCHEDULE_START_COLUMN,
    normalise_date,
)

KNOWN_RANKS = {
    "REC",
    "PTE",
    "LCP",
    "CPL",
    "CFC",
    "3SG",
    "2SG",
    "1SG",
    "SSG",
    "MSG",
    "3WO",
    "2WO",
    "1WO",
    "MWO",
    "SWO",
    "ME1",
    "ME2",
    "ME3",
    "ME4",
    "ME5",
    "ME6",
    "ME7",
    "ME8",
}

NON_PERSONNEL_LABELS = {
    "DM",
    "CS1",
    "CS2",
    "CS/B",
    "SB1",
    "SB2",
    "AE",
    "RESERVE",
    "TEMP COVER ROW",
    "COMBAT POINTS",
    "TOTAL POINTS",
    "NUMBER OF DM",
    "NUMBER OF SB",
    "NUMBER OF CS",
    "NUMBER OF AE",
    "TOTAL DUTY TEAM STRENGTH",
}


# Fixed personnel blocks in the current Scheduling Roster workbook template.
# These are intentionally bounded so synchronisation never inserts/deletes rows
# and therefore cannot shift the summary/formula sections below them.
PERSONNEL_BLOCKS = {
    "PT": range(5, 42),   # rows 5-41 inclusive
    "RH": range(52, 72),  # rows 52-71 inclusive
}

def canonical_person_name(value: object) -> str:
    """
    Return a standardised personnel name for matching Supabase records
    against names in the Excel roster.

    Examples:
        "CFC GERALD TAN "        -> "GERALD TAN"
        "LCP JARED JUAY (BCF)"   -> "JARED JUAY"
        "LCP GOH SONG YEE, ETHAN" -> "GOH SONG YEE ETHAN"
    """
    import re

    normalised = normalise_text(value)

    if not normalised:
        return ""

    # Remove trailing annotations such as "(BCF)".
    normalised = re.sub(
        r"\s*\([^)]*\)\s*$",
        "",
        normalised,
    )

    # Treat commas as spacing rather than part of the name.
    normalised = normalised.replace(",", " ")

    # Collapse repeated spaces.
    normalised = " ".join(normalised.split())

    parts = normalised.split()

    # Remove the rank prefix.
    if len(parts) >= 2 and parts[0] in KNOWN_RANKS:
        normalised = " ".join(parts[1:])

    return normalised

def roster_worksheet_name(year: int, month: int) -> str:
    """
    Return the expected roster worksheet name.

    Example:
        August 2026 -> Aug-2026 Roster
    """
    return date(year, month, 1).strftime("%b-%Y Roster")


def assignment_cell_value(assignment: Assignment) -> str:
    """
    Convert an internal role name into the value written into the roster.

    Examples:
        PT DM  -> DM
        RH SB1 -> SB1
        PT CS/B -> CS/B
    """
    role = assignment.role.strip()

    for prefix in ("PT ", "RH "):
        if role.upper().startswith(prefix):
            return role[len(prefix):].strip()

    return role


def find_date_columns(
    worksheet: Worksheet,
    *,
    year: int,
    month: int,
) -> dict[date, int]:
    """
    Map every date in the target month to its roster column.

    The roster stores the first date as a real Excel date while later
    date cells may contain formulas. Once the first date is found, the
    remaining columns are mapped sequentially.
    """
    from calendar import monthrange

    first_day = date(year, month, 1)
    days_in_month = monthrange(year, month)[1]

    first_date_column: int | None = None

    for column_number in range(
        SCHEDULE_START_COLUMN,
        worksheet.max_column + 1,
    ):
        cell_value = worksheet.cell(
            row=DATE_ROW,
            column=column_number,
        ).value

        cell_date = normalise_date(cell_value)

        if cell_date == first_day:
            first_date_column = column_number
            break

    if first_date_column is None:
        return {}

    return {
        date(year, month, day_number): (
            first_date_column + day_number - 1
        )
        for day_number in range(1, days_in_month + 1)
    }



def find_personnel_rows(
    worksheet: Worksheet,
) -> dict[str, int]:
    """
    Map personnel names (ignoring rank prefixes) to worksheet rows.
    """
    personnel_rows: dict[str, int] = {}

    for row_number in range(
        PERSONNEL_START_ROW,
        worksheet.max_row + 1,
    ):
        raw_name = worksheet.cell(
            row=row_number,
            column=PERSONNEL_COLUMN,
        ).value

        canonical_name = canonical_person_name(raw_name)

        if not canonical_name:
            continue

        if canonical_name in NON_PERSONNEL_LABELS:
            continue

        personnel_rows[canonical_name] = row_number

    return personnel_rows


def personnel_display_name(person: Person) -> str:
    """
    Return the value written into the Excel personnel-name column.

    Supabase is the source of truth for the person's full name and rank.
    """
    name = normalise_text(person.name)
    rank = normalise_text(person.rank)

    if not name:
        return ""

    parts = name.split()

    if parts and parts[0] in KNOWN_RANKS:
        return name

    if rank:
        return f"{rank} {name}"

    return name


def _person_is_in_target_month(
    person: Person,
    *,
    year: int,
    month: int,
) -> bool:
    """
    Include personnel who are active for at least part of the target month.

    leaving_date is treated as the first date the person is no longer usable.
    """
    if not person.is_active:
        return False

    month_start = date(year, month, 1)

    if (
        person.leaving_date is not None
        and person.leaving_date <= month_start
    ):
        return False

    return True


def _looks_like_same_person(
    excel_name: str,
    supabase_name: str,
) -> bool:
    """
    Conservative fallback matcher for legacy Excel names.

    Exact canonical matches are preferred. This fallback only handles:
    - one name being a clear prefix of the other; or
    - a very high character-similarity typo.

    It is used only when the match is unique within the same centre.
    """
    from difflib import SequenceMatcher

    excel_name = canonical_person_name(excel_name)
    supabase_name = canonical_person_name(supabase_name)

    if not excel_name or not supabase_name:
        return False

    if excel_name == supabase_name:
        return True

    excel_parts = excel_name.split()
    supabase_parts = supabase_name.split()

    shorter = excel_parts if len(excel_parts) <= len(supabase_parts) else supabase_parts
    longer = supabase_parts if len(excel_parts) <= len(supabase_parts) else excel_parts

    if (
        len(shorter) >= 2
        and longer[:len(shorter)] == shorter
    ):
        return True

    similarity = SequenceMatcher(
        None,
        excel_name,
        supabase_name,
    ).ratio()

    return similarity >= 0.92


def _clear_personnel_schedule_cells(
    worksheet: Worksheet,
    *,
    row_number: int,
    date_columns: dict[date, int],
) -> None:
    """
    Clear only the date-grid cells when a row is reassigned to another person.

    This prevents leave/duty values belonging to the old person from being
    inherited by the new Supabase person while preserving row formatting and
    summary formulas outside the date grid.
    """
    for column_number in date_columns.values():
        worksheet.cell(
            row=row_number,
            column=column_number,
        ).value = None


def sync_personnel_rows(
    worksheet: Worksheet,
    *,
    personnel: list[Person],
    year: int,
    month: int,
    date_columns: dict[date, int],
) -> dict[str, int]:
    """
    Synchronise PT/RH personnel names from Supabase into the fixed Excel blocks.

    Existing exact/unique legacy matches retain their rows so any legitimate
    template availability entries remain attached to the same person.

    Rows belonging to personnel no longer in the target month's Supabase pool
    are reused for unmatched/new personnel, and their date-grid cells are
    cleared before reuse.

    No rows are inserted or deleted.
    """
    target_people = [
        person
        for person in personnel
        if _person_is_in_target_month(
            person,
            year=year,
            month=month,
        )
    ]

    by_centre: dict[str, list[Person]] = {
        "PT": [],
        "RH": [],
    }

    for person in target_people:
        centre = normalise_text(person.centre)

        if centre not in by_centre:
            raise ValueError(
                f"Unsupported personnel centre for export: "
                f"{person.name} ({person.centre!r})"
            )

        by_centre[centre].append(person)

    final_rows: dict[str, int] = {}

    for centre, row_range in PERSONNEL_BLOCKS.items():
        people = by_centre[centre]
        available_rows = list(row_range)

        if len(people) > len(available_rows):
            raise ValueError(
                f"{centre} personnel count ({len(people)}) exceeds "
                f"the Excel template capacity ({len(available_rows)})."
            )

        existing_by_row = {
            row_number: canonical_person_name(
                worksheet.cell(
                    row=row_number,
                    column=PERSONNEL_COLUMN,
                ).value
            )
            for row_number in available_rows
        }

        unmatched_people = list(people)
        matched_rows: set[int] = set()

        # Pass 1: exact canonical match.
        for person in list(unmatched_people):
            person_key = canonical_person_name(
                person.name
            )

            exact_rows = [
                row_number
                for row_number, existing_key
                in existing_by_row.items()
                if (
                    row_number not in matched_rows
                    and existing_key == person_key
                )
            ]

            if len(exact_rows) == 1:
                row_number = exact_rows[0]
                matched_rows.add(row_number)
                final_rows[person_key] = row_number
                worksheet.cell(
                    row=row_number,
                    column=PERSONNEL_COLUMN,
                ).value = personnel_display_name(person)
                unmatched_people.remove(person)

        # Pass 2: conservative unique legacy-name match.
        for person in list(unmatched_people):
            candidate_rows = [
                row_number
                for row_number, existing_key
                in existing_by_row.items()
                if (
                    row_number not in matched_rows
                    and _looks_like_same_person(
                        existing_key,
                        person.name,
                    )
                )
            ]

            if len(candidate_rows) == 1:
                row_number = candidate_rows[0]
                matched_rows.add(row_number)
                person_key = canonical_person_name(
                    person.name
                )
                final_rows[person_key] = row_number
                worksheet.cell(
                    row=row_number,
                    column=PERSONNEL_COLUMN,
                ).value = personnel_display_name(person)
                unmatched_people.remove(person)

        # Reuse rows that no longer belong to a matched current person.
        reusable_rows = [
            row_number
            for row_number in available_rows
            if row_number not in matched_rows
        ]

        for person, row_number in zip(
            unmatched_people,
            reusable_rows,
            strict=False,
        ):
            _clear_personnel_schedule_cells(
                worksheet,
                row_number=row_number,
                date_columns=date_columns,
            )

            worksheet.cell(
                row=row_number,
                column=PERSONNEL_COLUMN,
            ).value = personnel_display_name(person)

            final_rows[
                canonical_person_name(person.name)
            ] = row_number

            matched_rows.add(row_number)

        # Clear unused personnel-name slots and their date-grid data.
        used_keys = {
            canonical_person_name(person.name)
            for person in people
        }

        for row_number in available_rows:
            current_key = canonical_person_name(
                worksheet.cell(
                    row=row_number,
                    column=PERSONNEL_COLUMN,
                ).value
            )

            if (
                row_number not in matched_rows
                or current_key not in used_keys
            ):
                _clear_personnel_schedule_cells(
                    worksheet,
                    row_number=row_number,
                    date_columns=date_columns,
                )
                worksheet.cell(
                    row=row_number,
                    column=PERSONNEL_COLUMN,
                ).value = None

    return final_rows

def copy_cell_style(
    source_cell,
    target_cell,
) -> None:
    """
    Copy formatting from one roster cell to another.

    This is useful when a blank template cell has lost formatting.
    """
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)

    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format

    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)

    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)


def export_schedule(
    *,
    template_path: Path,
    output_path: Path,
    schedule: Schedule,
    year: int,
    month: int,
    personnel: list[Person] | None = None,
    worksheet_name: str | None = None,
) -> Path:
    """
    Write generated assignments into a copy of the roster workbook.

    The original workbook is not modified.
    """
    if not template_path.exists():
        raise FileNotFoundError(
            f"Roster template was not found: {template_path}"
        )

    selected_worksheet = (
        worksheet_name
        or roster_worksheet_name(year, month)
    )

    workbook = openpyxl.load_workbook(template_path)

    try:
        if selected_worksheet not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet '{selected_worksheet}' was not found. "
                f"Available worksheets: {workbook.sheetnames}"
            )

        worksheet = workbook[selected_worksheet]

        date_columns = find_date_columns(
            worksheet,
            year=year,
            month=month,
        )

        if not date_columns:
            raise ValueError(
                f"No dates for {year}-{month:02d} were found "
                f"in worksheet '{selected_worksheet}'."
            )

        if personnel is not None:
            personnel_rows = sync_personnel_rows(
                worksheet,
                personnel=personnel,
                year=year,
                month=month,
                date_columns=date_columns,
            )
        else:
            personnel_rows = find_personnel_rows(
                worksheet
            )

        if not personnel_rows:
            raise ValueError(
                f"No personnel rows were found in "
                f"worksheet '{selected_worksheet}'."
            )

        missing_people: set[str] = set()
        missing_dates: set[date] = set()
        written_assignments = 0

        for assignment in schedule.assignments:
            if (
                assignment.duty_date.year != year
                or assignment.duty_date.month != month
            ):
                continue

            canonical_name = canonical_person_name(
                assignment.person_name
            )

            row_number = personnel_rows.get(canonical_name)
            column_number = date_columns.get(
                assignment.duty_date
            )

            if row_number is None:
                missing_people.add(assignment.person_name)
                continue

            if column_number is None:
                missing_dates.add(assignment.duty_date)
                continue

            target_cell = worksheet.cell(
                row=row_number,
                column=column_number,
            )

            new_value = assignment_cell_value(assignment)

            existing_value = (
                str(target_cell.value).strip()
                if target_cell.value not in (None, "")
                else ""
            )

            # Existing roster duty codes may be replaced by the newly generated
            # assignment. Availability and leave codes remain protected.
            replaceable_duty_codes = {
                "DM",
                "CS1",
                "CS2",
                "CS/B",
                "SB1",
                "SB2",
                "AE",
                "EM",
                "OUTFIELD RESERVE",
            }

            protected_availability_codes = {
                "AL",
                "ORD",
                "HL",
                "MC",
                "CCL",
                "OIL",
                "OFF",
                "MA",
                "AM",
                "PM",
            }

            normalised_existing = existing_value.upper()

            if normalised_existing in protected_availability_codes:
                raise ValueError(
                    "Cannot assign duty over an availability entry: "
                    f"{assignment.person_name}, "
                    f"{assignment.duty_date.isoformat()}, "
                    f"existing value={existing_value!r}, "
                    f"new value={new_value!r}"
                )

            if (
                existing_value
                and normalised_existing not in replaceable_duty_codes
            ):
                raise ValueError(
                    "Cannot safely overwrite an unrecognised roster entry: "
                    f"{assignment.person_name}, "
                    f"{assignment.duty_date.isoformat()}, "
                    f"existing value={existing_value!r}, "
                    f"new value={new_value!r}"
                )

            target_cell.value = new_value

            written_assignments += 1

        if missing_people:
            missing_names = ", ".join(
                sorted(missing_people)
            )
            raise ValueError(
                "Some assigned personnel could not be found "
                f"in the roster worksheet: {missing_names}"
            )

        if missing_dates:
            missing_date_text = ", ".join(
                roster_date.isoformat()
                for roster_date in sorted(missing_dates)
            )
            raise ValueError(
                "Some assignment dates could not be found "
                f"in the roster worksheet: {missing_date_text}"
            )

        if written_assignments == 0:
            raise ValueError(
                "No assignments were written to the roster."
            )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(output_path)

    finally:
        workbook.close()

    return output_path
