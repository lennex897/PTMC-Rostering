from __future__ import annotations

from pathlib import Path

import openpyxl

import re
from dataclasses import dataclass
from datetime import date

from roster_engine.models import (
    Assignment,
    Person,
    Schedule,
)
from roster_engine.personnel import (
    load_personnel,
    normalise_text,
)
from roster_engine.requirements import (
    overnight_points_for_date,
)
from roster_engine.roster_grid import parse_roster_grid


HISTORICAL_DUTY_ROLES = {
    "DM",
    "CS1",
    "CS2",
    "CS/B",
    "SB1",
    "SB2",
    "AE",
}

MONTH_NAMES = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


ROSTER_SHEET_PATTERN = re.compile(
    r"^\s*"
    r"(?P<month>[A-Za-z]{3})"
    r"[- ]"
    r"(?P<year>\d{4})"
    r"[- ]"
    r"Roster"
    r"\s*$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class HistoricalWorksheet:
    worksheet_name: str
    year: int
    month: int

    @property
    def month_start(self) -> date:
        return date(
            self.year,
            self.month,
            1,
        )

def normalise_historical_role(value: object) -> str | None:
    """
    Return a recognised historical duty role.

    Leave codes, deployment descriptions, reserve entries,
    formulas, names and summary values are ignored.
    """
    role = normalise_text(value)

    if role not in HISTORICAL_DUTY_ROLES:
        return None

    return role


def historical_role_name(
    centre: str,
    workbook_role: str,
) -> str:
    """
    Convert a workbook role such as DM or SB1 into the engine's
    centre-prefixed role format, such as PT DM or RH SB1.
    """
    return f"{normalise_text(centre)} {workbook_role}"

def parse_roster_worksheet_name(
    worksheet_name: str,
) -> HistoricalWorksheet | None:
    """
    Parse worksheet names such as:

    May-2026-Roster
    Jun-2026 Roster
    """
    match = ROSTER_SHEET_PATTERN.match(
        worksheet_name
    )

    if match is None:
        return None

    month_text = (
        match.group("month")
        .strip()
        .upper()
    )

    month = MONTH_NAMES.get(month_text)

    if month is None:
        return None

    return HistoricalWorksheet(
        worksheet_name=worksheet_name,
        year=int(match.group("year")),
        month=month,
    )


def find_historical_roster_sheets(
    workbook_path: Path,
    target_year: int,
    target_month: int,
    maximum_months: int | None = 3,
) -> list[str]:
    """
    Find roster worksheets before the target month.

    maximum_months:
        3    -> latest three available historical roster sheets
        None -> every available roster sheet before the target
    """
    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Roster workbook not found: {workbook_path}"
        )

    workbook = openpyxl.load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    try:
        target_start = date(
            target_year,
            target_month,
            1,
        )

        historical_sheets: list[
            HistoricalWorksheet
        ] = []

        for worksheet_name in workbook.sheetnames:
            parsed_sheet = (
                parse_roster_worksheet_name(
                    worksheet_name
                )
            )

            if parsed_sheet is None:
                continue

            if parsed_sheet.month_start >= target_start:
                continue

            historical_sheets.append(
                parsed_sheet
            )

        historical_sheets.sort(
            key=lambda item: (
                item.year,
                item.month,
            )
        )

        if maximum_months is not None:
            historical_sheets = (
                historical_sheets[
                    -maximum_months:
                ]
            )

        return [
            item.worksheet_name
            for item in historical_sheets
        ]

    finally:
        workbook.close()


def load_historical_schedules(
    workbook_path: Path,
    worksheet_names: list[str],
    personnel: list[Person] | None = None,
) -> Schedule:
    """
    Load and merge assignments from multiple historical
    roster worksheets.
    """
    if personnel is None:
        personnel = load_personnel(
            workbook_path
        )

    combined_schedule = Schedule()

    for worksheet_name in worksheet_names:
        worksheet_schedule = (
            load_historical_schedule(
                workbook_path=workbook_path,
                worksheet_name=worksheet_name,
                personnel=personnel,
            )
        )

        combined_schedule.assignments.extend(
            worksheet_schedule.assignments
        )

        combined_schedule.warnings.extend(
            worksheet_schedule.warnings
        )

    combined_schedule.assignments.sort(
        key=lambda assignment: (
            assignment.duty_date,
            assignment.person_name,
            assignment.role,
        )
    )

    return combined_schedule

def load_historical_schedule(
    workbook_path: Path,
    worksheet_name: str,
    personnel: list[Person] | None = None,
) -> Schedule:
    """
    Load historical duty assignments from a roster worksheet.

    Only recognised duty codes are imported:

    DM, CS1, CS2, CS/B, SB1, SB2 and AE.

    Personnel are matched against the authoritative Personnel
    worksheet, preventing summary tables from being interpreted
    as personnel rows.
    """
    if personnel is None:
        personnel = load_personnel(workbook_path)

    grid = parse_roster_grid(
        workbook_path=workbook_path,
        worksheet_name=worksheet_name,
    )

    people_by_name = {
        normalise_text(person.name): person
        for person in personnel
    }

    day_by_column = {
        roster_day.column_number: roster_day
        for roster_day in grid.days
    }

    schedule = Schedule()

    workbook = openpyxl.load_workbook(
        workbook_path,
        read_only=False,
        data_only=True,
    )

    try:
        worksheet = workbook[worksheet_name]

        first_date_column = min(day_by_column)
        last_date_column = max(day_by_column)

        for personnel_row in grid.personnel_rows:
            normalised_label = normalise_text(
                personnel_row.label
            )

            person = people_by_name.get(normalised_label)

            # Ignore summary labels and any names not found in
            # the authoritative Personnel worksheet.
            if person is None:
                continue

            for row in worksheet.iter_rows(
                min_row=personnel_row.row_number,
                max_row=personnel_row.row_number,
                min_col=first_date_column,
                max_col=last_date_column,
                values_only=True,
            ):
                for offset, raw_value in enumerate(row):
                    workbook_role = normalise_historical_role(
                        raw_value
                    )

                    if workbook_role is None:
                        continue

                    column_number = (
                        first_date_column + offset
                    )

                    roster_day = day_by_column.get(
                        column_number
                    )

                    if roster_day is None:
                        continue

                    schedule.add_assignment(
                        Assignment(
                            duty_date=roster_day.roster_date,
                            role=historical_role_name(
                                centre=person.centre,
                                workbook_role=workbook_role,
                            ),
                            centre=person.centre,
                            person_name=person.name,
                            points=overnight_points_for_date(
                                roster_day.roster_date
                            ),
                            is_overnight=True,
                        )
                    )

        return schedule

    finally:
        workbook.close()
