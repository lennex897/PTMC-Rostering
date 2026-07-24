from datetime import datetime, timedelta, date
from pathlib import Path

import openpyxl

from roster_engine.models import Person


EXCEL_EPOCH = datetime(1899, 12, 30)


def normalise_text(value: object) -> str:
    """Convert a cell value into a clean uppercase string."""
    if value is None:
        return ""

    return " ".join(str(value).strip().upper().split())


def parse_excel_date(value: object) -> date | None:
    """Convert Excel dates into Python dates."""
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    try:
        return (
            EXCEL_EPOCH
            + timedelta(days=float(value))
        ).date()
    except (TypeError, ValueError):
        return None


def extract_rank(name: str) -> str:
    """Extract the first word as the person's rank."""
    if not name:
        return ""

    return name.split()[0]


def load_personnel(
    workbook_path: Path,
) -> list[Person]:
    """
    Legacy workbook personnel loader.

    This loader is retained for historical/reference workbook support.
    Live personnel data for the application should come from Supabase.

    The workbook does not contain the newer Supabase-managed fields
    (service_type, is_cover_fit, or role eligibility), so those remain
    unset/defaulted here.
    """

    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Scheduling workbook not found: {workbook_path}"
        )

    workbook = openpyxl.load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    try:
        if "Personnel" not in workbook.sheetnames:
            raise ValueError(
                "The scheduling workbook does not contain a 'Personnel' worksheet."
            )

        worksheet = workbook["Personnel"]

        personnel: list[Person] = []

        # PT Personnel — Columns A:F
        for row_number in range(3, worksheet.max_row + 1):
            name = normalise_text(
                worksheet.cell(row_number, 1).value
            )

            if not name:
                continue

            ampt_status = normalise_text(
                worksheet.cell(row_number, 3).value
            )

            department = normalise_text(
                worksheet.cell(row_number, 4).value
            )

            leaving_date = parse_excel_date(
                worksheet.cell(row_number, 5).value
            )

            personnel.append(
                Person(
                    name=name,
                    rank=extract_rank(name),
                    centre="PT",
                    department=department or "UNSPECIFIED",
                    ampt_status=ampt_status,
                    leaving_date=leaving_date,
                )
            )

        # RH Personnel — Columns I:N
        for row_number in range(3, worksheet.max_row + 1):
            name = normalise_text(
                worksheet.cell(row_number, 9).value
            )

            if not name:
                continue

            ampt_status = normalise_text(
                worksheet.cell(row_number, 11).value
            )

            department = normalise_text(
                worksheet.cell(row_number, 12).value
            )

            leaving_date = parse_excel_date(
                worksheet.cell(row_number, 13).value
            )

            personnel.append(
                Person(
                    name=name,
                    rank=extract_rank(name),
                    centre="RH",
                    department=department or "UNSPECIFIED",
                    ampt_status=ampt_status,
                    leaving_date=leaving_date,
                )
            )

        return personnel

    finally:
        workbook.close()
