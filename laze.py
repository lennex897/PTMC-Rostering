import openpyxl

wb = openpyxl.load_workbook("reference/Scheduling Roster 2026.xlsx", data_only=True)
ws = wb["Aug-2026 Roster"]

for r in range(5, ws.max_row + 1):
    v = ws.cell(r, 2).value
    if v:
        print(v)