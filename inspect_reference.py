from pathlib import Path

import openpyxl


APP_ROOT = Path(__file__).resolve().parent

workbook_path = (
    APP_ROOT
    / "reference"
    / "Scheduling Roster 2026.xlsx"
)

workbook = openpyxl.load_workbook(
    workbook_path,
    read_only=True,
    data_only=True,
)

try:
    for worksheet in workbook.worksheets:
        print(f"\n### {worksheet.title}")
        print(
            f"Rows: {worksheet.max_row}, "
            f"Columns: {worksheet.max_column}"
        )

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(10, worksheet.max_row),
                min_col=1,
                max_col=min(20, worksheet.max_column),
                values_only=True,
            ),
            start=1,
        ):
            values = [
                f"C{column_number}={value!r}"
                for column_number, value in enumerate(row, start=1)
                if value is not None
            ]

            if values:
                print(f"Row {row_number}: " + " | ".join(values))

finally:
    workbook.close()